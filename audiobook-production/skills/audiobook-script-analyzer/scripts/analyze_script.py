#!/usr/bin/env python3
"""
Audiobook Script Analyzer
Analyzes PDF or Word documents to extract chapter structure, word counts, and page counts.
Generates a spreadsheet with chapter information and file naming conventions.
"""

import re
import sys
import json
import argparse
from pathlib import Path
from dataclasses import dataclass, field, asdict
from typing import Optional

# PDF handling
try:
    import pdfplumber
    HAS_PDFPLUMBER = True
except ImportError:
    HAS_PDFPLUMBER = False

# Word document handling
try:
    from docx import Document
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False

# Excel output
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


@dataclass
class Chapter:
    """Represents a chapter/section in the audiobook."""
    number: int  # Sequential number for file naming (001, 002, etc.)
    title: str   # Chapter title as it appears (e.g., "Chapter One", "Prologue")
    chapter_type: str  # prologue, chapter, interlude, epilogue, etc.
    start_page: int
    end_page: int
    page_count: int = 0
    word_count: int = 0
    file_name: str = ""
    
    def __post_init__(self):
        self.page_count = self.end_page - self.start_page + 1


@dataclass
class AnalysisResult:
    """Results from analyzing an audiobook script."""
    book_title: str
    book_short_name: str
    total_word_count: int
    total_pages: int
    chapter_count: int
    chapters: list
    front_matter_pages: int = 0
    back_matter_pages: int = 0
    front_matter_end_page: int = 0
    back_matter_start_page: int = 0


# Chapter detection patterns
CHAPTER_PATTERNS = [
    # Prologue variations
    (r'^(?:PROLOGUE|Prologue)(?:\s*[-–—:]?\s*)?$', 'prologue'),
    # Epilogue variations  
    (r'^(?:EPILOGUE|Epilogue)(?:\s*[-–—:]?\s*)?$', 'epilogue'),
    # Interlude variations
    (r'^(?:INTERLUDE|Interlude)(?:\s*[-–—:]?\s*)?$', 'interlude'),
    # Chapter with written numbers (Chapter One, CHAPTER TWO, etc.)
    (r'^(?:CHAPTER|Chapter)\s+([A-Za-z]+(?:\s*[-–—]\s*[A-Za-z]+)?)(?:\s*[-–—:]?\s*)?$', 'chapter'),
    # Chapter with numerals (Chapter 1, CHAPTER 12, etc.)
    (r'^(?:CHAPTER|Chapter)\s+(\d+)(?:\s*[-–—:]?\s*)?$', 'chapter'),
    # Just numbers on their own line (1, 12, etc.) - less reliable, use with caution
    (r'^(\d{1,3})$', 'chapter_number_only'),
    # Part markers (Part One, PART 1, etc.)
    (r'^(?:PART|Part)\s+([A-Za-z]+|\d+)(?:\s*[-–—:]?\s*)?$', 'part'),
    # Act markers
    (r'^(?:ACT|Act)\s+([A-Za-z]+|\d+)(?:\s*[-–—:]?\s*)?$', 'act'),
    # Introduction
    (r'^(?:INTRODUCTION|Introduction)(?:\s*[-–—:]?\s*)?$', 'introduction'),
    # Afterword
    (r'^(?:AFTERWORD|Afterword)(?:\s*[-–—:]?\s*)?$', 'afterword'),
]

# Front matter indicators (to help identify where content starts)
FRONT_MATTER_KEYWORDS = [
    'copyright', 'title page', 'dedication', 'table of contents',
    'acknowledgments', 'acknowledgements', 'foreword', 'preface',
    'all rights reserved', 'isbn', 'published by'
]

# Back matter indicators
BACK_MATTER_KEYWORDS = [
    'about the author', 'also by', 'other books', 'appendix',
    'glossary', 'index', 'bibliography', 'notes', 'afterword',
    'author bio', 'connect with', 'follow the author', 'litrpg',
    'gamelit', 'portal books', 'blackstone', 'audible'
]

# Written number to integer mapping
WRITTEN_NUMBERS = {
    'one': 1, 'two': 2, 'three': 3, 'four': 4, 'five': 5,
    'six': 6, 'seven': 7, 'eight': 8, 'nine': 9, 'ten': 10,
    'eleven': 11, 'twelve': 12, 'thirteen': 13, 'fourteen': 14, 'fifteen': 15,
    'sixteen': 16, 'seventeen': 17, 'eighteen': 18, 'nineteen': 19, 'twenty': 20,
    'twenty-one': 21, 'twenty-two': 22, 'twenty-three': 23, 'twenty-four': 24,
    'twenty-five': 25, 'twenty-six': 26, 'twenty-seven': 27, 'twenty-eight': 28,
    'twenty-nine': 29, 'thirty': 30, 'thirty-one': 31, 'thirty-two': 32,
    'thirty-three': 33, 'thirty-four': 34, 'thirty-five': 35, 'thirty-six': 36,
    'thirty-seven': 37, 'thirty-eight': 38, 'thirty-nine': 39, 'forty': 40,
    'forty-one': 41, 'forty-two': 42, 'forty-three': 43, 'forty-four': 44,
    'forty-five': 45, 'forty-six': 46, 'forty-seven': 47, 'forty-eight': 48,
    'forty-nine': 49, 'fifty': 50, 'fifty-one': 51, 'fifty-two': 52,
    'fifty-three': 53, 'fifty-four': 54, 'fifty-five': 55, 'fifty-six': 56,
    'fifty-seven': 57, 'fifty-eight': 58, 'fifty-nine': 59, 'sixty': 60,
    'sixty-one': 61, 'sixty-two': 62, 'sixty-three': 63, 'sixty-four': 64,
    'sixty-five': 65, 'sixty-six': 66, 'sixty-seven': 67, 'sixty-eight': 68,
    'sixty-nine': 69, 'seventy': 70, 'seventy-one': 71, 'seventy-two': 72,
    'seventy-three': 73, 'seventy-four': 74, 'seventy-five': 75,
    'seventy-six': 76, 'seventy-seven': 77, 'seventy-eight': 78,
    'seventy-nine': 79, 'eighty': 80, 'eighty-one': 81, 'eighty-two': 82,
    'eighty-three': 83, 'eighty-four': 84, 'eighty-five': 85, 'eighty-six': 86,
    'eighty-seven': 87, 'eighty-eight': 88, 'eighty-nine': 89, 'ninety': 90,
    'ninety-one': 91, 'ninety-two': 92, 'ninety-three': 93, 'ninety-four': 94,
    'ninety-five': 95, 'ninety-six': 96, 'ninety-seven': 97, 'ninety-eight': 98,
    'ninety-nine': 99, 'one hundred': 100,
}


def extract_text_from_pdf(filepath: str) -> list[dict]:
    """Extract text from PDF, returning list of {page: int, text: str, word_count: int}."""
    if not HAS_PDFPLUMBER:
        raise ImportError("pdfplumber is required for PDF analysis. Install with: pip install pdfplumber")
    
    pages = []
    with pdfplumber.open(filepath) as pdf:
        for i, page in enumerate(pdf.pages, 1):
            text = page.extract_text() or ""
            words = len(text.split())
            pages.append({
                'page': i,
                'text': text,
                'word_count': words,
                'lines': [line.strip() for line in text.split('\n') if line.strip()]
            })
    return pages


def extract_text_from_docx(filepath: str) -> list[dict]:
    """Extract text from Word document, simulating pages based on content."""
    if not HAS_DOCX:
        raise ImportError("python-docx is required for Word document analysis. Install with: pip install python-docx")
    
    doc = Document(filepath)
    
    # Word docs don't have natural page breaks we can always detect
    # We'll estimate ~300 words per page and group paragraphs
    all_paragraphs = []
    for para in doc.paragraphs:
        text = para.text.strip()
        if text:
            all_paragraphs.append(text)
    
    # Group into approximate pages
    pages = []
    current_page_text = []
    current_word_count = 0
    page_num = 1
    words_per_page = 300  # Approximate
    
    for para in all_paragraphs:
        para_words = len(para.split())
        current_page_text.append(para)
        current_word_count += para_words
        
        if current_word_count >= words_per_page:
            full_text = '\n'.join(current_page_text)
            pages.append({
                'page': page_num,
                'text': full_text,
                'word_count': current_word_count,
                'lines': [line.strip() for line in full_text.split('\n') if line.strip()]
            })
            current_page_text = []
            current_word_count = 0
            page_num += 1
    
    # Don't forget the last page
    if current_page_text:
        full_text = '\n'.join(current_page_text)
        pages.append({
            'page': page_num,
            'text': full_text,
            'word_count': len(full_text.split()),
            'lines': [line.strip() for line in full_text.split('\n') if line.strip()]
        })
    
    return pages


def detect_chapter(line: str) -> Optional[tuple[str, str]]:
    """
    Check if a line is a chapter header.
    Returns (title, chapter_type) or None.
    """
    line = line.strip()
    
    # Skip very long lines - chapter headers are usually short
    if len(line) > 50:
        return None
    
    for pattern, chapter_type in CHAPTER_PATTERNS:
        match = re.match(pattern, line, re.IGNORECASE)
        if match:
            # For chapter_number_only, be more conservative
            if chapter_type == 'chapter_number_only':
                # Only accept if it's a reasonable chapter number
                num = int(match.group(1))
                if num > 150 or num < 1:
                    continue
            return (line, chapter_type)
    
    return None


def is_front_matter(text: str) -> bool:
    """Check if text appears to be front matter."""
    text_lower = text.lower()
    return any(keyword in text_lower for keyword in FRONT_MATTER_KEYWORDS)


def is_back_matter(text: str) -> bool:
    """Check if text appears to be back matter."""
    text_lower = text.lower()
    return any(keyword in text_lower for keyword in BACK_MATTER_KEYWORDS)


def find_chapters(pages: list[dict]) -> list[dict]:
    """
    Find all chapters in the document.
    Returns list of {title, chapter_type, start_page, line_index}.
    """
    chapters = []
    
    for page_data in pages:
        page_num = page_data['page']
        
        for line_idx, line in enumerate(page_data['lines']):
            result = detect_chapter(line)
            if result:
                title, chapter_type = result
                chapters.append({
                    'title': title,
                    'chapter_type': chapter_type,
                    'start_page': page_num,
                    'line_index': line_idx
                })
    
    return chapters


def analyze_script(filepath: str, book_short_name: str) -> AnalysisResult:
    """
    Main analysis function.
    """
    filepath = Path(filepath)
    
    # Extract text based on file type
    suffix = filepath.suffix.lower()
    if suffix == '.pdf':
        pages = extract_text_from_pdf(str(filepath))
    elif suffix in ['.docx', '.doc']:
        pages = extract_text_from_docx(str(filepath))
    elif suffix == '.txt':
        # Simple text file handling - preserve line structure
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            all_lines = [line.strip() for line in f.readlines()]
        
        # Group lines into pages (approximately 30 lines per page)
        lines_per_page = 30
        pages = []
        for i in range(0, len(all_lines), lines_per_page):
            page_lines = all_lines[i:i+lines_per_page]
            page_text = '\n'.join(page_lines)
            word_count = len(page_text.split())
            pages.append({
                'page': (i // lines_per_page) + 1,
                'text': page_text,
                'word_count': word_count,
                'lines': [line for line in page_lines if line]
            })
    else:
        raise ValueError(f"Unsupported file type: {suffix}")
    
    total_pages = len(pages)
    
    # Find all chapters
    raw_chapters = find_chapters(pages)
    
    if not raw_chapters:
        raise ValueError("No chapters detected in the document. Please check the file format.")
    
    # Determine front matter end (page before first chapter)
    first_chapter_page = raw_chapters[0]['start_page']
    front_matter_end = first_chapter_page - 1 if first_chapter_page > 1 else 0
    
    # Determine back matter start (page after last chapter ends)
    # We'll calculate this after determining chapter end pages
    
    # Build chapter objects with end pages
    chapters = []
    for i, ch in enumerate(raw_chapters):
        # End page is page before next chapter starts, or last page
        if i + 1 < len(raw_chapters):
            end_page = raw_chapters[i + 1]['start_page'] - 1
        else:
            end_page = total_pages
        
        chapters.append({
            'title': ch['title'],
            'chapter_type': ch['chapter_type'],
            'start_page': ch['start_page'],
            'end_page': end_page
        })
    
    # Check for back matter after last chapter
    # Look at the last few pages for back matter indicators
    last_chapter = chapters[-1]
    back_matter_start = total_pages + 1  # Default: no back matter
    
    # Scan from last chapter's last pages for back matter
    for page_data in reversed(pages[last_chapter['start_page']-1:]):
        if is_back_matter(page_data['text']):
            # Found back matter indicator, adjust last chapter end
            back_matter_start = page_data['page']
        else:
            break
    
    # Adjust last chapter end page if we found back matter
    if back_matter_start <= total_pages:
        chapters[-1]['end_page'] = back_matter_start - 1
    
    # Calculate word counts for each chapter
    for ch in chapters:
        word_count = 0
        for page_data in pages[ch['start_page']-1:ch['end_page']]:
            word_count += page_data['word_count']
        ch['word_count'] = word_count
    
    # Create Chapter objects with file names
    chapter_objects = []
    for i, ch in enumerate(chapters, 1):
        file_number = f"{i:03d}"
        # Clean title for filename (remove special characters, keep the header only)
        clean_title = ch['title'].replace(':', '').replace('—', '-').replace('–', '-')
        clean_title = re.sub(r'[<>:"/\\|?*]', '', clean_title)
        
        file_name = f"{file_number}_{book_short_name}_{clean_title}"
        
        chapter_obj = Chapter(
            number=i,
            title=ch['title'],
            chapter_type=ch['chapter_type'],
            start_page=ch['start_page'],
            end_page=ch['end_page'],
            word_count=ch['word_count'],
            file_name=file_name
        )
        chapter_objects.append(chapter_obj)
    
    # Calculate totals (excluding front/back matter)
    story_pages = pages[front_matter_end:back_matter_start-1] if back_matter_start <= total_pages else pages[front_matter_end:]
    total_word_count = sum(p['word_count'] for p in story_pages)
    total_story_pages = len(story_pages)
    
    # Calculate front/back matter page counts
    front_matter_pages = front_matter_end
    back_matter_pages = total_pages - back_matter_start + 1 if back_matter_start <= total_pages else 0
    
    return AnalysisResult(
        book_title=filepath.stem,
        book_short_name=book_short_name,
        total_word_count=total_word_count,
        total_pages=total_story_pages,
        chapter_count=len(chapter_objects),
        chapters=[asdict(ch) for ch in chapter_objects],
        front_matter_pages=front_matter_pages,
        back_matter_pages=back_matter_pages,
        front_matter_end_page=front_matter_end,
        back_matter_start_page=back_matter_start if back_matter_start <= total_pages else 0
    )


def generate_spreadsheet(result: AnalysisResult, output_path: str):
    """Generate Excel spreadsheet with chapter information."""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required for Excel output. Install with: pip install openpyxl")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Chapter Analysis"
    
    # Styles
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Summary section at top
    ws['A1'] = "Book Analysis Summary"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    
    summary_data = [
        ("Book Title:", result.book_title),
        ("Short Name:", result.book_short_name),
        ("Total Word Count:", f"{result.total_word_count:,}"),
        ("Total Story Pages:", result.total_pages),
        ("Number of Chapters:", result.chapter_count),
        ("Front Matter Pages:", result.front_matter_pages),
        ("Back Matter Pages:", result.back_matter_pages),
    ]
    
    for i, (label, value) in enumerate(summary_data, 3):
        ws[f'A{i}'] = label
        ws[f'A{i}'].font = Font(bold=True)
        ws[f'B{i}'] = value
    
    # Chapter table header
    header_row = 12
    headers = ["#", "Chapter Title", "File Name", "Pages", "Start Page", "End Page", "Word Count"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Chapter data
    for i, chapter in enumerate(result.chapters, 1):
        row = header_row + i
        ws.cell(row=row, column=1, value=chapter['number']).border = thin_border
        ws.cell(row=row, column=2, value=chapter['title']).border = thin_border
        ws.cell(row=row, column=3, value=chapter['file_name']).border = thin_border
        ws.cell(row=row, column=4, value=chapter['page_count']).border = thin_border
        ws.cell(row=row, column=5, value=chapter['start_page']).border = thin_border
        ws.cell(row=row, column=6, value=chapter['end_page']).border = thin_border
        ws.cell(row=row, column=7, value=chapter['word_count']).border = thin_border
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12
    
    wb.save(output_path)
    return output_path


def main():
    parser = argparse.ArgumentParser(description='Analyze audiobook script for chapter structure')
    parser.add_argument('filepath', help='Path to PDF or Word document')
    parser.add_argument('--short-name', '-s', required=True, help='Short name for file naming (e.g., DotF, Exlian4)')
    parser.add_argument('--output', '-o', help='Output Excel file path')
    parser.add_argument('--json', action='store_true', help='Output JSON instead of Excel')
    
    args = parser.parse_args()
    
    # Analyze the script
    result = analyze_script(args.filepath, args.short_name)
    
    if args.json:
        # Output JSON
        print(json.dumps(asdict(result) if hasattr(result, '__dataclass_fields__') else {
            'book_title': result.book_title,
            'book_short_name': result.book_short_name,
            'total_word_count': result.total_word_count,
            'total_pages': result.total_pages,
            'chapter_count': result.chapter_count,
            'chapters': result.chapters,
            'front_matter_pages': result.front_matter_pages,
            'back_matter_pages': result.back_matter_pages,
        }, indent=2))
    else:
        # Generate Excel
        output_path = args.output or f"{Path(args.filepath).stem}_analysis.xlsx"
        generate_spreadsheet(result, output_path)
        print(f"Analysis complete. Spreadsheet saved to: {output_path}")
        
        # Also print summary
        print(f"\n=== Summary ===")
        print(f"Word Count: {result.total_word_count:,}")
        print(f"Story Pages: {result.total_pages}")
        print(f"Chapters: {result.chapter_count}")


if __name__ == '__main__':
    main()
